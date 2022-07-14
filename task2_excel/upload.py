from sqlalchemy import create_engine, Column, Integer, Text
from sqlalchemy.orm import declarative_base, Session
from openpyxl import load_workbook
from sys import argv

EXEL_FILE = "названия точек.xlsm"
EXEL_SHEET = "Лист1"
EXEL_COL_WITH_ID = 1
EXEL_COL_WITH_NAME = 2
SQL_URL = "sqlite:///local_test"

if argv[1]:
    EXEL_FILE = argv[1]

BaseTable = declarative_base()


class Endpoint(BaseTable):
    __tablename__ = "endpoint_names"
    id = Column(Integer, primary_key=True)
    endpoint_id = Column(Integer)
    name = Column(Text)


engine = create_engine(SQL_URL)

BaseTable.metadata.create_all(engine)
session = Session(engine)

exel_file = load_workbook(EXEL_FILE)
sheet = exel_file[EXEL_SHEET]


def get_row(row: int):
    return sheet.cell(row=row, column=EXEL_COL_WITH_ID).value, sheet.cell(row=row, column=EXEL_COL_WITH_NAME).value


row = 2
while (row_val := get_row(row))[0]:
    id, name = row_val
    session.add(Endpoint(endpoint_id=id, name=name))

    row += 1
session.commit()
